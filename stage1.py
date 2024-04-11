import os
import json
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def process_xml_file(xml_file, ws, row):
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()

        namespace = {"ns": "urn:cbr-440P:oper4:v4.00..3"}

        attributes_data = {}

        for item in root.findall(".//*"):
            attributes = item.attrib
            for attr_name, attr_value in attributes.items():
                if attr_name not in attributes_data:
                    attributes_data[attr_name] = []
                attributes_data[attr_name].append(attr_value)

        col = 1
        for attr_name, attr_values in attributes_data.items():
            col_letter = get_column_letter(col)
            ws[f"{col_letter}{row}"] = attr_name
            for i, attr_value in enumerate(attr_values):
                ws[f"{col_letter}{row+i+1}"] = attr_value
            col += 1

    except ET.ParseError as e:
        print(
            f"Ошибка при обработке файла {xml_file}: {str(e)}")


def xml_folder_to_xlsx(xml_folder, xlsx_folder):
    xml_files = [f for f in os.listdir(xml_folder) if f.endswith(".xml")]
    xml_files.sort()

    unique_structures = {}
    xlsx_files = {}

    for xml_file in xml_files:
        xml_path = os.path.join(xml_folder, xml_file)

        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()

            attributes_data = {}

            for item in root.findall(".//*"):
                attributes = item.attrib
                for attr_name, attr_value in attributes.items():
                    if attr_name not in attributes_data:
                        attributes_data[attr_name] = []
                    attributes_data[attr_name].append(attr_value)

            attributes_key = json.dumps(attributes_data, sort_keys=True)
            if attributes_key not in unique_structures:
                unique_structures[attributes_key] = []

            unique_structures[attributes_key].append(xml_file)

        except ET.ParseError as e:
            print(
                f"Ошибка при разборе XML-файла {xml_file}: {str(e)}")
            continue

    for attributes_key, xml_files in unique_structures.items():
        current_xlsx_file = Workbook()
        ws = current_xlsx_file.active

        attributes_data = json.loads(attributes_key)

        header_row = list(attributes_data.keys())
        ws.append(header_row)

        row = 2

        for xml_file in xml_files:
            xml_path = os.path.join(xml_folder, xml_file)
            process_xml_file(xml_path, ws, row)
            row += len(attributes_data) + 1

        xlsx_file_name = f"merged_{len(xlsx_files)+1}.xlsx"
        xlsx_file_path = os.path.join(xlsx_folder, xlsx_file_name)
        current_xlsx_file.save(xlsx_file_path)
        xlsx_files[attributes_key] = xlsx_file_path

    return xlsx_files


# Пример использования:
xml_folder = r'C:\Users\kyana\Desktop\test\Камелот'
xlsx_folder = r'C:\Users\kyana\Desktop\test\Камелот xlsx'

result = xml_folder_to_xlsx(xml_folder, xlsx_folder)
